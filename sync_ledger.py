from __future__ import annotations

import os
import time
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

SOURCE_PATH = Path(
    os.environ.get(
        "SHAREPOINT_LEDGER_PATH",
        r"C:\Users\박인영(InyeongPark)\OneDrive - 뉴로핏 주식회사\R&D\00. 연구개발과제 관리대장\뉴로핏_연구과제_통합관리.xlsx",
    )
)
OUT_DIR = Path(__file__).resolve().parent / "ledger"
OUT_PATH = OUT_DIR / "총괄표_원본데이터.xlsx"
POLL_SECONDS = 5.0

# 과제 시트(예: "101", "2")의 연차별 요약표는 K열부터 시작 (단계-연차/시작일/종료일/정부지원금 등)
WIDE_START_COL = 11
# 세목별 상세표(단계-연차/세목/.../예산(현금)/...)에 등장하는 세목 화이트리스트.
# 이 목록에 없는 항목이 나오면 상세 집행내역 등 다른 구간에 진입한 것으로 보고 읽기를 중단한다.
COST_ITEMS = ["인건비", "연구활동비", "연구재료비", "연구시설장비비", "연구수당", "간접비", "매출", "타기관"]

WONBON_COLUMNS = [
    "시트명", "사업명", "연차수", "단계-연차", "시작일", "종료일",
    "정부지원금 (현금)", "지방비 (현금)", "민간부담금 (현금)", "민간부담금 (현물)", "합계",
    "인건비", "연구활동비", "연구재료비", "연구시설장비비", "연구수당", "간접비", "매출",
    "타기관", "기술료",
]


def _norm(v: object) -> str:
    if v is None:
        return ""
    return str(v).replace("\n", " ").strip()


def _read_wide_headers(ws) -> dict:
    headers = {}
    for c in range(WIDE_START_COL, WIDE_START_COL + 20):
        key = _norm(ws.cell(row=1, column=c).value)
        if key:
            headers[key] = c
    return headers


def _find_long_header_row(ws) -> Optional[int]:
    for r in range(1, 60):
        a = _norm(ws.cell(row=r, column=1).value)
        b = _norm(ws.cell(row=r, column=2).value)
        if a == "단계-연차" and b == "세목":
            return r
    return None


def read_project_rows(ws, sheet_name: str) -> list[dict]:
    """과제 시트(연차별 요약표 K:Y + 세목별 상세표 A:G)를 원본데이터와 같은 스키마의 행으로 재구성.

    과거 '원본데이터' 시트는 이 두 표를 INDIRECT/INDEX/MATCH 수식으로 참조해 만들어졌다.
    """
    wide_headers = _read_wide_headers(ws)
    year_col = wide_headers.get("단계-연차", WIDE_START_COL)
    start_col = wide_headers.get("시작일")

    year_rows: list[tuple] = []
    r = 2
    blank_streak = 0
    while r < 200:
        label = ws.cell(row=r, column=year_col).value
        label_n = _norm(label)
        if not label_n:
            blank_streak += 1
            if blank_streak > 2:
                break
        elif label_n == "총합":
            break
        else:
            blank_streak = 0
            # 단계-연차 라벨은 향후 연차를 위해 미리 채워져 있을 수 있음 →
            # 실제로 시작일이 입력된(=확정된) 연차만 포함
            has_start_date = start_col and ws.cell(row=r, column=start_col).value is not None
            if has_start_date:
                year_rows.append((r, label))
        r += 1

    long_header_row = _find_long_header_row(ws)
    allowed_items = set(COST_ITEMS) | {"합계"}
    cost_map: dict = {}
    if long_header_row:
        r = long_header_row + 1
        blank_streak = 0
        while r < long_header_row + 400:
            a_n = _norm(ws.cell(row=r, column=1).value)
            b = _norm(ws.cell(row=r, column=2).value)
            if not a_n and not b:
                blank_streak += 1
                if blank_streak > 3:
                    break
            else:
                blank_streak = 0
                if b and b not in allowed_items:
                    break
                if b and b != "합계":
                    val = ws.cell(row=r, column=4).value  # 예산(현금)
                    cost_map[(a_n, b)] = val or 0
            r += 1

    def wv(row, key):
        col = wide_headers.get(key)
        return ws.cell(row=row, column=col).value if col else None

    business_name = ws.cell(row=4, column=2).value  # B4 = 사업명

    rows = []
    for idx, (r, label) in enumerate(year_rows, start=1):
        label_n = _norm(label)
        gov = wv(r, "정부지원금 (현금)") or 0
        local = wv(r, "지방비 (현금)") or 0
        priv_cash = wv(r, "민간부담금 (현금)") or 0
        priv_kind = wv(r, "민간부담금 (현물)") or 0

        rows.append({
            "시트명": sheet_name,
            "사업명": business_name,
            "연차수": idx,
            "단계-연차": label,
            "시작일": wv(r, "시작일"),
            "종료일": wv(r, "종료일"),
            "정부지원금 (현금)": gov,
            "지방비 (현금)": local,
            "민간부담금 (현금)": priv_cash,
            "민간부담금 (현물)": priv_kind,
            "합계": (gov or 0) + (local or 0) + (priv_cash or 0) + (priv_kind or 0),
            "인건비": cost_map.get((label_n, "인건비"), 0) or 0,
            "연구활동비": cost_map.get((label_n, "연구활동비"), 0) or 0,
            "연구재료비": cost_map.get((label_n, "연구재료비"), 0) or 0,
            "연구시설장비비": cost_map.get((label_n, "연구시설장비비"), 0) or 0,
            "연구수당": cost_map.get((label_n, "연구수당"), 0) or 0,
            "간접비": cost_map.get((label_n, "간접비"), 0) or 0,
            "매출": cost_map.get((label_n, "매출"), 0) or 0,
            "타기관": wv(r, "타기관") or 0,
            "기술료": wv(r, "기술료") or 0,
        })
    return rows


def build_wonbon_dataframe(wb, summary_df: pd.DataFrame) -> pd.DataFrame:
    """총괄표의 번호 중 실제 과제 시트가 존재하는 것만 순회해 원본데이터를 재구성."""
    numbers = summary_df["번호"].dropna().tolist()

    all_rows: list[dict] = []
    for n in numbers:
        sheet_name = str(int(n)) if float(n).is_integer() else str(n)
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        try:
            all_rows.extend(read_project_rows(ws, sheet_name))
        except Exception as exc:
            print(f"[LEDGER-SYNC] 시트 '{sheet_name}' 읽기 실패, 건너뜀: {exc}")

    return pd.DataFrame(all_rows, columns=WONBON_COLUMNS)


def export_ledger() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    summary_df = pd.read_excel(SOURCE_PATH, sheet_name="총괄표", header=0)
    wb = openpyxl.load_workbook(SOURCE_PATH, data_only=True)
    try:
        wonbon_df = build_wonbon_dataframe(wb, summary_df)
    finally:
        wb.close()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="총괄표", index=False)
        wonbon_df.to_excel(writer, sheet_name="원본데이터", index=False)


def main() -> None:
    print(f"[LEDGER-SYNC] 감시 시작: {SOURCE_PATH}")
    print(f"[LEDGER-SYNC] polling: {POLL_SECONDS}초, 출력: {OUT_PATH}")

    last_mtime: float | None = None
    while True:
        if not SOURCE_PATH.exists():
            print(f"[LEDGER-SYNC] 원본 파일을 찾을 수 없습니다: {SOURCE_PATH}")
        else:
            mtime = SOURCE_PATH.stat().st_mtime
            if mtime != last_mtime:
                try:
                    export_ledger()
                    print(f"[LEDGER-SYNC] 갱신 완료 -> {OUT_PATH}")
                except Exception as exc:
                    print(f"[LEDGER-SYNC] 갱신 실패: {exc}")
                last_mtime = mtime

        time.sleep(POLL_SECONDS)


if __name__ == "__main__":
    main()
