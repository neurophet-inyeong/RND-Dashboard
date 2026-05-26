import json
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
PROJECTS_DIR = ROOT / "projects"
GENERATED_DIR = ROOT / "generated"
JSON_OUTPUT = GENERATED_DIR / "project-data.json"
JS_OUTPUT = GENERATED_DIR / "project-data.js"

OVERVIEW_KEYS = {
    "유형": "type",
    "부처": "ministry",
    "전문기관": "agency",
    "사업명": "programName",
    "과제명": "projectName",
    "과제번호": "projectCode",
    "주관/공동(책임자)": "owner",
    "총 수행기간": "totalPeriod",
    "당해기간": "currentPeriod",
    "당해기간(개월)": "currentPeriodMonths",
}

FIXED_METRIC_FIELDS = [
    "metricName",
    "verification",
    "unit",
    "weight",
    "organization",
]


def clean_text(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 4)
    if isinstance(value, int):
        return value
    text = str(value).strip()
    return text or None


def normalize_row(values):
    return [clean_text(value) for value in values]


def first_nonempty_after(values, start_index):
    for value in values[start_index:]:
        cleaned = clean_text(value)
        if cleaned is not None:
            return cleaned
    return None


def parse_overview(sheet):
    meta = {}
    raw_rows = []

    for row in sheet.iter_rows(values_only=True):
        values = normalize_row(row)
        if not any(value is not None for value in values):
            continue
        key = values[0]
        if key is None:
            continue
        value = first_nonempty_after(values, 1)
        raw_rows.append({"label": key, "value": value})
        normalized_key = OVERVIEW_KEYS.get(key)
        if normalized_key:
            meta[normalized_key] = value

    return {
        "meta": meta,
        "rows": raw_rows,
    }


def parse_budget(sheet):
    headers = normalize_row(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    items = []
    summary = None

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = normalize_row(row)
        if not any(value is not None for value in values):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            record[header] = values[index] if index < len(values) else None
        phase = record.get("단계-연차")
        total_budget = record.get("총 사업비")
        if phase == "총합":
            summary = record
        elif phase is None and total_budget in (0, None):
            continue
        else:
            items.append(record)

    return {
        "headers": [header for header in headers if header is not None],
        "items": items,
        "summary": summary,
    }


def milestone_status(item):
    def parse_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None

    today = date.today()
    plan_start = parse_date(item.get("planStart"))
    actual_start = parse_date(item.get("actualStart"))
    actual_end = parse_date(item.get("actualEnd"))

    if actual_end is not None:
        return "completed"
    if actual_start is not None:
        return "in-progress"
    if plan_start is not None and plan_start >= today:
        return "scheduled"
    if plan_start is not None and plan_start < today:
        return "delayed"
    return "scheduled"


def parse_milestones(sheet):
    items = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        values = normalize_row(row)
        if not any(value is not None for value in values):
            continue
        if values[0] is None:
            continue

        item = {
            "no": values[0],
            "content": values[1],
            "phase": values[2],
            "planStart": values[3],
            "planEnd": values[4],
            "actualStart": values[5],
            "actualEnd": values[6],
        }
        item["status"] = milestone_status(item)
        items.append(item)

    return {
        "items": items,
    }


def normalize_metric_group(group_name):
    group_name = group_name or ""
    if "목표" in group_name:
        return "target"
    if "달성치" in group_name or group_name.strip() == "달성" or "달성 " in group_name:
        return "actual"
    if "달성률" in group_name:
        return "rate"
    return None


def parse_metrics(sheet):
    top_headers = normalize_row(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    sub_headers = normalize_row(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))

    column_map = []
    current_group = None
    # Fixed columns are A-D (지표명/확인지표/단위/가중치), dynamic metric columns start at E.
    for index in range(4, sheet.max_column):
        header_group = top_headers[index] if index < len(top_headers) else None
        if header_group is not None:
            current_group = header_group

        period = sub_headers[index] if index < len(sub_headers) else None
        metric_group = normalize_metric_group(current_group)

        if period is None or metric_group is None:
            continue

        column_map.append(
            {
                "index": index,
                "period": period,
                "metricGroup": metric_group,
            }
        )

    items = []
    periods = []
    seen_periods = set()

    for mapping in column_map:
        period = mapping["period"]
        if period not in seen_periods:
            periods.append(period)
            seen_periods.add(period)

    for row in sheet.iter_rows(min_row=3, values_only=True):
        values = normalize_row(row)
        if not any(value is not None for value in values):
            continue
        if values[0] is None:
            continue

        item = {
            FIXED_METRIC_FIELDS[0]: values[0],
            FIXED_METRIC_FIELDS[1]: values[1],
            FIXED_METRIC_FIELDS[2]: values[2],
            FIXED_METRIC_FIELDS[3]: values[3],
            # The 5th column is the first period target, not organization metadata.
            FIXED_METRIC_FIELDS[4]: None,
            "periods": [],
        }
        by_period = {}

        for mapping in column_map:
            value = values[mapping["index"]] if mapping["index"] < len(values) else None
            period = mapping["period"]
            if period not in by_period:
                by_period[period] = {
                    "period": period,
                    "target": None,
                    "actual": None,
                    "rate": None,
                }
            by_period[period][mapping["metricGroup"]] = value

        item["periods"] = [by_period[period] for period in periods if period in by_period]
        items.append(item)

    return {
        "periods": periods,
        "items": items,
    }


def parse_metrics_detail(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {"headers": [], "items": []}

    raw_headers = normalize_row(rows[0])
    max_column_index = len(raw_headers) - 1

    normalized_rows = [normalize_row(row) for row in rows[1:]]
    for values in normalized_rows:
        for index, value in enumerate(values):
            if value is not None:
                max_column_index = max(max_column_index, index)

    if max_column_index < 0:
        return {"headers": [], "items": []}

    headers = []
    for index in range(max_column_index + 1):
        header = raw_headers[index] if index < len(raw_headers) else None
        if header is None:
            if index == 4:
                header = "상세"
            else:
                header = f"항목{index + 1}"
        headers.append(header)

    items = []
    for values in normalized_rows:
        record = {}
        has_value = False
        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else None
            record[header] = value
            if value is not None:
                has_value = True
        if has_value:
            items.append(record)

    return {
        "headers": headers,
        "items": items,
    }


def build_project_payload(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)

    overview = parse_overview(workbook["개요"]) if "개요" in workbook.sheetnames else {"meta": {}, "rows": []}
    budget = parse_budget(workbook["사업비"]) if "사업비" in workbook.sheetnames else {"headers": [], "items": [], "summary": None}
    milestones = parse_milestones(workbook["마일스톤"]) if "마일스톤" in workbook.sheetnames else {"items": []}
    metrics = parse_metrics(workbook["정량지표"]) if "정량지표" in workbook.sheetnames else {"periods": [], "items": []}
    metrics_detail = parse_metrics_detail(workbook["정량지표상세"]) if "정량지표상세" in workbook.sheetnames else {"headers": [], "items": []}

    meta = overview["meta"]
    project_code = meta.get("projectCode") or workbook_path.stem

    return {
        "id": project_code,
        "fileName": workbook_path.name,
        "sheetNames": workbook.sheetnames,
        "overview": overview,
        "budget": budget,
        "milestones": milestones,
        "metrics": metrics,
        "metricsDetail": metrics_detail,
    }


def build_dataset():
    files = sorted(
        path for path in PROJECTS_DIR.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    projects = [build_project_payload(path) for path in files]

    return {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sourceDirectory": str(PROJECTS_DIR),
        "projectCount": len(projects),
        "projects": projects,
    }


def main():
    GENERATED_DIR.mkdir(exist_ok=True)
    dataset = build_dataset()
    serialized = json.dumps(dataset, ensure_ascii=False, indent=2)
    JSON_OUTPUT.write_text(serialized, encoding="utf-8")
    JS_OUTPUT.write_text(f"window.PROJECT_DATA = {serialized};\n", encoding="utf-8")
    print(f"Wrote {JSON_OUTPUT}")
    print(f"Wrote {JS_OUTPUT}")


if __name__ == "__main__":
    main()